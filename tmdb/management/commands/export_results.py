from django.core.management.base import BaseCommand
from openpyxl import *
import os, sys
from tmdb.models import *



class ExportResults():
	def __init__(self):
		self.spreadsheet = load_workbook("/Users/andreaguatemala/Desktop/Results2016template.xlsx") # Change path to wherever results are to be saved

		self.sparringSheet =  self.spreadsheet.get_sheet_by_name("Sparring")
		self.outputSheet = self.spreadsheet.get_sheet_by_name("Output_4")

		self.sparring_cell = [4, 4]
		self.output_cell = [4, 3]

		self.download_results()

	def download_results(self):

		self.tournament = Tournament.objects.first()
		self.divisions = TournamentDivision.objects.filter(tournament = self.tournament)

		# for division in self.divisions:
		for division in self.divisions:
			skill = DivisionLevelEnum.get(division.division.skill_level).label
			sex = SexEnum.get(division.division.sex).label

			final = TeamMatch.objects.filter(division = division, round_num = 0)[0]
			semis = TeamMatch.objects.filter(division = division, round_num = 1)
			quarters = TeamMatch.objects.filter(division = division, round_num = 2)

			if sex == 'Male':
				if skill == 'A':
					self.sparring_cell = [4, 4]
					self.output_cell = [4, 3]
				elif skill == 'B':
					self.sparring_cell = [13, 4]
					self.output_cell = [63, 3]
				elif skill == 'C':
					self.sparring_cell = [22, 4]
					self.output_cell = [122, 3]

			if sex == 'Female':
				if skill == 'A':
					self.sparring_cell = [4, 9]
					self.output_cell = [4, 8]
				elif skill == 'B':
					self.sparring_cell = [13, 9]
					self.output_cell = [63, 8]
				elif skill == 'C':
					self.sparring_cell = [22, 9]
					self.output_cell = [122, 8]

			# Finals Results

			if final.winning_team.team == final.red_team.team:
				losing_team = final.blue_team
			else:
				losing_team = final.red_team
			winning_school = final.winning_team.team.school
			winning_teamNum = str(skill) + str(final.winning_team.team.number)
			winning_competitiors = [final.winning_team.lightweight, final.winning_team.middleweight, final.winning_team.heavyweight, final.winning_team.alternate1, final.winning_team.alternate2]
			losing_school = losing_team.team.school
			losing_teamNum = str(skill) + str(losing_team.team.number)
			losing_competitiors = [losing_team.lightweight, losing_team.middleweight, losing_team.heavyweight, losing_team.alternate1, losing_team.alternate2]

			## Sparring Sheet

			self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(winning_school))
			self.sparring_cell[1] +=1
			self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(winning_teamNum))
			self.sparring_cell[0] +=1
			self.sparring_cell[1] -=1
			self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_school))
			self.sparring_cell[1] +=1
			self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_teamNum))
			self.sparring_cell[0] +=1
			self.sparring_cell[1] -=1

			## Output sheet

			self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(winning_school))
			self.output_cell[1] += 1
			self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(winning_teamNum))
			self.output_cell[1] -= 1
			self.output_cell[0] += 1

			for comp in winning_competitiors:
				if comp is not None:
					self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = comp.name)
				self.output_cell[0] +=1

			self.output_cell[0] += 1
			self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_school))
			self.output_cell[1] += 1
			self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_teamNum))
			self.output_cell[1] -= 1
			self.output_cell[0] += 1

			for comp in losing_competitiors:
				if comp is not None:
					self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = comp.name)
				self.output_cell[0] +=1

			# Semis results

			for semi in semis:
				if semi.winning_team.team == semi.red_team.team:
					losing_team = semi.blue_team
				else:
					losing_team = semi.red_team
				losing_school = losing_team.team.school
				losing_teamNum = str(skill) + str(losing_team.team.number)
				losing_competitiors = [losing_team.lightweight, losing_team.middleweight, losing_team.heavyweight, losing_team.alternate1, losing_team.alternate2]

				## Sparring Sheet

				self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_school))
				self.sparring_cell[1] +=1
				self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_teamNum))
				self.sparring_cell[0] +=1
				self.sparring_cell[1] -=1

				#3 Output Sheet

				self.output_cell[0] += 1
				self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_school))
				self.output_cell[1] += 1
				self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_teamNum))
				self.output_cell[1] -= 1
				self.output_cell[0] += 1

				for comp in losing_competitiors:
					if comp is not None:
						self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = comp.name)
					self.output_cell[0] +=1

			# Quarters results

			for quarter in quarters:
				if quarter.winning_team.team == quarter.red_team.team:
					losing_team = quarter.blue_team
				else:
					losing_team = quarter.red_team
				losing_school = losing_team.team.school
				losing_teamNum = str(skill) + str(losing_team.team.number)
				losing_competitiors = [losing_team.lightweight, losing_team.middleweight, losing_team.heavyweight, losing_team.alternate1, losing_team.alternate2]

				## Sparring Sheet

				self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_school))
				self.sparring_cell[1] +=1
				self.sparringSheet.cell(row = self.sparring_cell[0], column = self.sparring_cell[1], value = str(losing_teamNum))
				self.sparring_cell[0] +=1
				self.sparring_cell[1] -=1

				## Output Sheet

				self.output_cell[0] += 1
				self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_school))
				self.output_cell[1] += 1
				self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = str(losing_teamNum))
				self.output_cell[1] -= 1
				self.output_cell[0] += 1

				for comp in losing_competitiors:
					if comp is not None:
						self.outputSheet.cell(row = self.output_cell[0], column = self.output_cell[1], value = comp.name)
					self.output_cell[0] +=1


		self.spreadsheet.save("/Users/andreaguatemala/Desktop/Results2016template.xlsx") # Change path to wherever results are to be saved

class Command(BaseCommand):
    help = 'Exports database results'

    def handle(self, *args, **options):
        ExportResults()

